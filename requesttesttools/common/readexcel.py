import openpyxl, re
import requests
from common.addlogs import add_logs
from common.readjsondata import read_json_data
from common.redfilepathini import ReadIni
from config.EXCELcells import ExcelCells


# 读取Excel内容的类
class ReadExcelData:

    def __init__(self):
        # 实例化ReadIni类，用于获取Excel的路径
        self.getdata = ReadIni()
        # 通过readexcel方法，获取excel文件的完整路径
        self.excelfile = self.getdata.readexcel()
        # 打开excel文件，准备进行读取
        self.workbook = openpyxl.load_workbook(self.excelfile)
        # 从配置文件中读取Sheet页的名字
        self.sheet = self.getdata.redexcelsheet()
        # 获取excel数据，并将其公有化
        self.exceldata = self.workbook[self.sheet]
        # 实例化ExcelCells类，用于获取表格单元格的常量，比如A列单元格
        self.excelcell = ExcelCells()
        # 将json参数文件提升为一个变量
        self.jsonfile = self.getdata.readjson()

    def write_excel_data(self, column, row, data):
        """写入excel，请在写入时关闭表格"""
        if data:
            self.exceldata["{}{}".format(column, row)].value = data
            self.workbook.save(self.excelfile)

    def read_excel_data(self, column, row):
        """
        读取Excel单元格内容的方法
        :param column: 列
        :param row: 行
        :return: 返回获取到的数据
        """
        return self.exceldata["{}{}".format(column, row)].value

    def read_excel_url(self, row):
        """
        获取url数据
        :param row:
        :return:
        """
        column = self.excelcell.CELL_REQUEST_URL
        url = self.read_excel_data(column, row)
        if "ksuperminiapp" in url:
            newurl = url.split("ksuperminiapp")[1]
            befurl = self.readu_r_l()
            url = befurl + "/ksuperminiapp" + newurl

        if url != None:
            try:
                if "ms_session" in url:
                    session = self.read_session()
                    session = re.subn(":", "%3A", session)
                    session = session[0]
                    # regular = re.compile("ms_session=(.+?)")
                    str1 = re.findall("&ms_session=(.+?)&", url)
                    # str1 = re.findall("ms_session=(.+?)", url)

                    url = re.subn(str1[0], session, url)

                    return url[0]
                else:
                    return url
            except:
                add_logs().error("URL中的session格式错误，无法更新到最新的session")
        else:
            add_logs().error("url为空，请检查数据")
            return False

    def read_case_id(self, row):
        """
        获取用例ID
        :param row:
        :return:
        """
        column = self.excelcell.CELL_TEST_ID
        return self.read_excel_data(column, row)

    def read_request_method(self, row):
        """
        获取请求方法
        :param row:
        :return:
        """
        column = self.excelcell.CELL_REQUEST_METHOD
        return self.read_excel_data(column, row)

    def read_params(self, row):
        """
        获取请求参数，这里一般指用于POST方法的参数（data）
        :param row:
        :return:
        """
        column = self.excelcell.CELL_REQUEST_PARAMS
        paramstitle = self.read_excel_data(column, row)
        parmas = read_json_data(self.jsonfile).get(paramstitle)
        if parmas != None:
            if "ms_session" in parmas:
                parmas["ms_session"] = self.read_session()
                return parmas
            else:
                return parmas
        else:
            return parmas

    def read_headers(self, row):
        """
        获取请求头，请求头一般可以指保存一个在json中，但是需要每天维护session参数
        :param row:
        :return:
        """
        column = self.excelcell.CELL_REQUEST_HEADERS
        headerstitle = self.read_excel_data(column, row)
        headers = read_json_data(self.jsonfile).get(headerstitle)
        # 将headers中的session更换为最新的
        if headers != None:
            if "ms_session" in headers:
                headers["ms_session"] = self.read_session()
                return headers
            else:
                return headers
        else:
            return headers

    def read_beforecaseid(self, row):
        """
        读取前置用例ID
        :param row:
        :return:
        """
        column = self.excelcell.CELL_BEFORECASEID
        return self.read_excel_data(column, row)

    def get_max_row(self):
        """
        获取最大行数
        :return:
        """
        return self.exceldata.max_row

    def get_max_column(self):
        """
        获取最大列数
        :return:
        """
        return self.exceldata.max_column

    def read_regular(self, row):
        """
        获取最大行数
        :param row:
        :return:
        """
        column = self.excelcell.CELL_REGULAR
        return self.read_excel_data(column, row)

    def read_dependent(self, row):
        """获取依赖字段，用于更新params"""
        column = self.excelcell.CELL_DEPENDENT
        return self.read_excel_data(column, row)

    def read_expect(self, row):
        """
        读取期望字段
        :param row:
        :return:
        """
        column = self.excelcell.CELL_EXPECT
        return self.read_excel_data(column, row)

    def read_is_excute(self, row):
        """
        获取是否执行该用例的控制字段
        :param row:
        :return:
        """
        column = self.excelcell.CELL_ISEXCUTE
        return self.read_excel_data(column, row)

    def read_casename(self, row):
        """
        读取用例名字，用例描述
        :param row:
        :return:
        """
        column = self.excelcell.CELL_CASENAME
        return self.read_excel_data(column, row)

    def read_session(self):
        """
        读取session，用于更换prams和header中的session
        :param row:
        :return:
        """
        column = self.excelcell.CELL_SESSION
        session = self.read_excel_data(column, 2)
        # session = self.get_Session()
        if session != None:
            return session

    # 参数化数据
    def paramsexpectdata(self):
        """
        对用例执行时进行参数化配置，在参数化中调用该方法，即可传入相关参数
        :return:
        """
        #返回的参数列表
        self.list1 = []
        #参数化excel文件和json文件
        for excelfile,jsonfile in self.getdata.get_excel_json_dict().items():
            curren_excel = excelfile
            curren_json = jsonfile
            self.excelfile = self.getdata.get_current_path(curren_excel)
            self.jsonfile = self.getdata.get_current_path(curren_json)
            #参数化sheet页
            for i in self.workbook.sheetnames:
                self.sheet = i
                self.exceldata=self.workbook[self.sheet]
                #参数化期望字段，和行数
                for row in range(2, self.get_max_row() + 1):
                    # 控制用例是否执行
                    if self.read_is_excute(row) != False:
                        expect = self.read_expect(row)
                        self.list1.append((row, expect,self.sheet, self.excelfile, self.jsonfile))
                    else:
                        add_logs().info("跳过该条用例：{}".format(self.read_casename(row)))
        return self.list1

    def readu_r_l(self):
        """读取测试环境"""
        column = self.excelcell.U_R_L
        u = self.read_excel_data(column, 2)
        if u:
            return u

    def read_dependent_param_type(self, row):
        """
        读取依赖参数的类型，默认字典的value，LIST->返回一个参数列表
        :param row:
        :return:
        """
        colunm = self.excelcell.DEPENDENT_PARAMS_TYPE
        type = self.read_excel_data(colunm, row)
        if type != None:
            return type

    def get_Session(self):
        """没用，废弃"""
        requests.packages.urllib3.disable_warnings()  # 避免https报错
        parmas = {"appid": "wxd5647f839d12c83a",
                  "clienttype": "wxapp",
                  "version": "3.0.0",
                  "mpkey": "mp_f0eecb40-23e0-11eb-8fd0-bbf1de6a814e",
                  "mininfo_id": "564",
                  "productcode": "ksuperminiapp"}
        url = self.readu_r_l() + "/ksuperminiapp/api/login/getSession"
        data = requests.get(params=parmas, url=url, verify=False)
        response = data.json()
        if response:
            res = response["data"]
            # return res
            column = self.excelcell.CELL_SESSION
            self.write_excel_data(column=column, row=2, data=res)


if __name__ == '__main__':
    # print(type(ReadExcelData().get_max_column))
    e = ReadExcelData()



    print(e.paramsexpectdata())
    pass
