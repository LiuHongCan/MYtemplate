import openpyxl
from APItest.APItestConfig.CellsTitle import ExcelCellTitle
from APItest.common.RediniPath import ReadIni
from APItest.common.readjsondata import ReadJsonData


class Readexceldata:
    """
    读取Excel中的数据类
    """
    def __init__(self):
        """
        初始化
        """
        # 获取Excel文件路径
        excelpath = ReadIni().Readexcelpath()
        # 载入Excel文件
        workbook = openpyxl.load_workbook(excelpath)
        # 获取sheet页的名字
        sheet = ReadIni().ReadSheet()
        # 获取sheet页中的内容
        self.exceldata = workbook[sheet]
        # 获取Excel中的列的编号;比如 A 对应第一列，加上row数据，组成Excel中的单元格[A1]
        self.exceltitle = ExcelCellTitle()

    def read_excelcell(self, column, row):
        """
        传入列，和行数，进行单元格数据读取
        :param column:
        :param row:
        :return:
        """
        return self.exceldata["{}{}".format(column, row)].value

    def read_testcaseid(self, row):
        """
        读取用例编号
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_TEST_CASE_ID
        return self.read_excelcell(column, row)

    def read_testcasetitle(self, row):
        """
        读取用例标题或者用例描述
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_TEST_CASE_TITLE
        return self.read_excelcell(column, row)

    def read_request_method(self, row):
        """
        获取请求方法
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_REQUEST_METHOD
        return self.read_excelcell(column, row)

    def read_request_url(self, row):
        """
        获取请求url
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_REQUEST_URL
        return self.read_excelcell(column, row)

    def read_case_execute_branch(self, row):
        """
        读取是否执行该用例的关键字
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_TEST_CASE_EXECUTE_BRANCH
        return self.read_excelcell(column, row)

    def read_front_case_id(self, row):
        """
        获取前置用例ID
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_FRONT_TEST_CASE_ID
        return self.read_excelcell(column, row)

    def read_relguar(self, row):
        """
        获取正则表达式
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_REGULAR
        return self.read_excelcell(column, row)

    def read_dependent_fields(self, row):
        """
        获取依赖字段，用于和正则表达式提取的字段组成字典
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_DEPENDENT_FIELDS
        return self.read_excelcell(column, row)

    def read_params(self, row):
        """
        读取要提交的参数
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_REQUEST_PARAMS
        paramsfield = self.read_excelcell(column, row)
        if paramsfield:
            return ReadJsonData()[paramsfield]

    def read_params_cell(self, row):
        column = self.exceltitle.CELL_REQUEST_PARAMS
        return self.read_excelcell(column, row)

    def get_cell_max_row(self):
        """
        获取当前页的最大行数
        :return: 
        """
        return self.exceldata.max_row
    
    def get_cell_max_column(self):
        return self.exceldata.max_column

    def read_cell_expect(self, row):
        """
        获取期望字段
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_EXPECT
        expect = self.read_excelcell(column, row)
        return expect

    def get_response_data_type(self, row):
        """
        获取指定返回的响应数据格式
        :param row:
        :return:
        """
        column = self.exceltitle.CELL_DATA_TYPE
        return self.read_excelcell(column, row)

    def paramsexpectdata(self):
        self.list1 = []
        for row in range(2, self.get_cell_max_row() + 1):
            expect = self.read_cell_expect(row)
            self.list1.append((row, expect))
        return self.list1

    # def read_front_testcase(self, row):
    #     column = self.exceltitle.CELL_FRONT_TEST_CASE_ID
    #     return self.read_excelcell(column, row)



if __name__ == '__main__':
    print(Readexceldata().paramsexpectdata())

